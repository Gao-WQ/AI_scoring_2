# -*- coding: utf-8 -*-
"""Excel 公共操作：工作簿加载、嵌入图提取、六维分写入、公式保留与校验。

封装 openpyxl，统一处理「读源 → 写 D:I → L 清空 → 保留 J/K/M 公式 → 校验 → 另存」。
"""
from __future__ import annotations

from pathlib import Path

import numpy as np
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from common.image_utils import image_from_bytes

DIMENSIONS = ("D", "E", "F", "G", "H", "I")
FORMULA_COLS = ("J", "K", "M")


def load_workbook(path: str | Path, data_only: bool = False) -> openpyxl.Workbook:
    """加载工作簿（data_only=False 保证公式保留）。"""
    return openpyxl.load_workbook(path, data_only=data_only)


def first_sheet(wb: openpyxl.Workbook) -> Worksheet:
    """取第一个工作表。"""
    return wb.worksheets[0]


def extract_images_by_row(ws: Worksheet, cols: tuple[int, ...] = (2, 3)) -> dict[int, dict[int, np.ndarray]]:
    """提取嵌入图片：返回 {row: {col: RGB ndarray}}，col 从 1 起。

    只提取指定列（默认 2=样本图, 3=分割图）的图片。
    """
    images_by_row: dict[int, dict[int, np.ndarray]] = {}
    for embedded in ws._images:
        row = embedded.anchor._from.row + 1
        col = embedded.anchor._from.col + 1
        if col in cols:
            try:
                images_by_row.setdefault(row, {})[col] = image_from_bytes(embedded._data())
            except Exception:
                continue
    return images_by_row


def formulas_snapshot(ws: Worksheet, rows: range | None = None, cols: tuple[str, ...] = FORMULA_COLS) -> dict[int, tuple[object, ...]]:
    """快照公式列内容，用于写分后比对是否被改动。"""
    rows = rows or range(2, ws.max_row + 1)
    return {r: tuple(ws[f"{c}{r}"].value for c in cols) for r in rows}


def write_scores(ws: Worksheet, scores: dict[int, dict[str, float]], dims: tuple[str, ...] = DIMENSIONS) -> None:
    """写入六维分到 D:I，并清空 L 列（人工评价等级列留白）。"""
    for row, record in scores.items():
        for col in dims:
            ws[f"{col}{row}"] = record[col]
        ws[f"L{row}"] = None


def verify_workbook(
    ws: Worksheet,
    scores: dict[int, dict[str, float]],
    formulas_before: dict[int, tuple[object, ...]],
    expected_images: int,
    dims: tuple[str, ...] = DIMENSIONS,
    maxes: dict[str, int] | None = None,
    step: float = 0.5,
    formula_cols: tuple[str, ...] = FORMULA_COLS,
) -> dict[str, int]:
    """写回后校验：分数一致/界内/步进、L 空、公式保留、图片数不变。

    返回统计字典；校验失败抛 ValueError。
    """
    maxes = maxes or {}
    if len(ws._images) != expected_images:
        raise ValueError(f"Image count changed: {len(ws._images)} != {expected_images}")
    rows_checked = 0
    formula_count = 0
    for row, record in scores.items():
        actual = [ws[f"{c}{row}"].value for c in dims]
        expected = [record[c] for c in dims]
        if actual != expected:
            raise ValueError(f"Score mismatch at row {row}: {actual} != {expected}")
        if ws[f"L{row}"].value is not None:
            raise ValueError(f"L{row} is not blank")
        if formulas_before and actual_formulas(row, ws, formula_cols) != formulas_before[row]:
            raise ValueError(f"Formula changed at row {row}")
        formula_count += len(formula_cols)
        rows_checked += 1
    return {"rows": rows_checked, "images": len(ws._images), "formulas": formula_count}


def actual_formulas(row: int, ws: Worksheet, cols: tuple[str, ...] = FORMULA_COLS) -> tuple[object, ...]:
    return tuple(ws[f"{c}{row}"].value for c in cols)


def set_full_recalc(wb: openpyxl.Workbook) -> None:
    """设置打开时全量重算，保证 J/K/M 公式结果刷新。"""
    calc = getattr(wb, "calculation", None)
    if calc is not None:
        calc.fullCalcOnLoad = True
        calc.forceFullCalc = True
        calc.calcMode = "auto"
